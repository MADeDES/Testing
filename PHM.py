import pandas as pd
import numpy as np
import pymongo
import openpyxl
from collections.abc import MutableMapping

FLOW = "Flow Property"
SYMP = "Symptom"
BOTH = "Both"
COMPONENT = "Component"
FAILURE = "Failure"
FM_LABELS = ['_FailureMode__uid', 'Diagnostic Group', 'Component', 'Flow Property', 'Failure']


def get_failure_mode_list(pt):
    """
    Get list of unique failure modes
    :return:
    """
    unique = {}
    fm = pt[COMPONENT] + " " + pt[FLOW] + " " + pt[FAILURE]
    fm_list = []
    for index, row in pt.iterrows():
        key = fm[index]
        if unique.get(key) is None:
            unique[key] = [index]
            fm_list.append(key)
        else:
            fm_list.append(key + "." + str(len(unique[key])))
            unique[key].append(index)

    return fm_list #pt[COMPONENT] + " " + pt[FLOW] + " " + pt[FAILURE]

def search_value_in_col_idx(ws, search_string, col_idx=1):
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx, None

def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row

def try_create_ws(filepath, sheet):
    wb = openpyxl.load_workbook(filepath, read_only=False)
    try:
        ws = wb[sheet]
    except KeyError:
        wb.create_sheet(sheet)
        ws = wb[sheet]
        # raise Exception(f"Sheet {sheet} does not exist")
    return wb, ws

def cursor_len(cursor):
    """
    Return cursor length
    :param cursor:
    :return:
    """
    i = 0
    for doc in cursor:
        i += 1
    return i

def fm_stats(pt):

    """
    fm stats are input to other equations that can be used to rapidly calculate statistics inlcuding:
        - Coverage
        - Detection
        - Criticality, cost etc... of uncovered failure modes
        - Criticality, cost etc... of undetected failure modes
    :param pt:
    :return:
    """

    num_fm = pt.shape[0]
    (unique, indexes, inv_indexes, counts) = np.unique(pt, axis=0, return_index=True, return_inverse=True, return_counts=True)
    return (unique, indexes, inv_indexes, counts, num_fm)

def calc_coverage_np2(counts, num_fm):
    return np.count_nonzero(counts == 1)/num_fm

def calc_coverage_np(pt):
    (unique, indexes, counts) = np.unique(pt, axis=0, return_index=True, return_counts=True)
    return np.count_nonzero(counts == 1)/pt.shape[0]

def calc_detection_np(pt):
    count = 0
    for row in pt:
        if np.abs(row).sum() == 0:
            count += 1

    return 1-count/pt.shape[0]

def obj_2_collection(obj, coll, key):
    """
    Send object to databse in dict- > JSON -> BSON
    :param coll:
    :return:
    """
    dict_for_db = obj.obj_2_dict()

    coll.delete_one({key: dict_for_db[key]})
    #if cursor_len(cursor) == 0:
        # listing["available"] = True # Is this listing still available
    try:
        coll.insert_one(dict_for_db)
    except:
        raise Exception("error")


def _flatten_dict_gen(d, parent_key, sep):
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, MutableMapping):
            yield from flatten_dict(v, new_key, sep=sep).items()
        else:
            yield new_key, v

def flatten_dict(d: MutableMapping, parent_key: str = '', sep: str = '.'):
    return dict(_flatten_dict_gen(d, parent_key, sep))


class SensorFactory:
    """
    Read in excel document, create python objects and send it to a JSON file for MongoDB, then delete the python objects
    """
    def __init__(self, path, sheetname, collection):
        df = pd.read_excel(path, sheetname, index_col=0)
        # assert that all columns of the sensor library are unique (unique uids)
        assert 1==1
        for column in df:
            sensor = Sensor(df[column].to_frame())
            obj_2_collection(sensor, collection, "_Sensor__uid")


class FmFactory:
    """
    Read in excel document, create python objects and send it to a JSON file for MongoDB, then delete the python objects
    """
    def __init__(self, path, sheetname, collection):
        self.df = pd.read_excel(path, sheetname, index_col=0)
        # assert that all columns of the sensor library are unique (unique uids)
        assert 1==1
        for column in self.df:
            fm = FailureMode(self.df[column].to_frame())
            obj_2_collection(fm, collection, "_FailureMode__uid")


class S_LocFactory:
    """
    Read in excel document, create python objects and send it to a JSON file for MongoDB, then delete the python objects
    """
    def __init__(self, path, sheetname, collection):
        df = pd.read_excel(path, sheetname, index_col=0)
        # assert that all columns of the sensor library are unique (unique uids)
        assert 1==1
        for column in df:
            s_loc = SenseLocation(df[column].to_frame())
            obj_2_collection(s_loc, collection, "_SenseLocation__uid")


# class GenomeNode:
#     """
#     Node of sensor set optimsation
#     """
#     def __init__(self):
#         self.genome = None # set of indexes that correspond to which sensor/location pairs are chosen for the GA
#         self.metrics = None # Table of df of all the relevant metrics form  the dataset


class SensorSolutionSpace:
    """
    Base Object Thate enables sensor set optimsation
    """
    def __init__(self):
        self.pt = None # Numpy Pt of
        self.loc_sensor_table = None # map s_locs to sensors (assert dimesions match pt)
        self.expl_loc_sensor_table = None # exploded table, one sensor/location pair per column
        self.fm_table = None

    # def calc_metrics(self, genome):
    #     """
    #
    #     :param genome: from GenomeNode
    #     :return: sensor set metrics to GenomeNode
    #     """
    #     metrics = None
    #     return metrics

def genetic_algorithm(sss, depth, epochs):
    """
    :param sss: SensorSolutionSpace
    :param depth: Depth of heap to be inspected
    :param epochs:
    :return:
    """
    #https://blog.paperspace.com/working-with-different-genetic-algorithm-representations-python/


class Sensor:
    """
    """
    def __init__(self, df=None):
        self.__uid = None  #
        self.name = None
        self.method_of_sensing = "On-line"
        self.automation = "Automatic"
        self.sensed_variables = None
        self.acquisition_cost = 0
        self.replacement_cost = 0
        self.operational_cost_ph = 0
        self.testing_cost_ph = 0
        self.false_alarm_cost = 0
        self.detection_cost = 0
        self.mttf = None
        self.mttr = None
        self.sampling_interval = None
        self.test_duration = 0
        self.personnel = []
        self.equipment = []
        self.dimensions = (0, 0, 0)
        self.volume = None
        self.weight = 0
        self.error_code = ""

        self.attr_dict = None
        #self.num_metrics = None

        self.metric_table = None

        if df is not None:
            self.dict_2_obj(df)

    def dict_2_obj(self, df):

        """
        Create object from mongodb database (dict) or from excel
        hardcode the creation of the item from the schema
        :param df:
        :return:
        """
        if isinstance(df, pd.DataFrame):
            key = list(df.keys())[0]
            self.__uid = key
            df = df[key]

            if not pd.isnull(df["personnel"]):
                self.equipment = df["equipment"].split(',')
            if not pd.isnull(df["equipment"]):
                self.equipment = df["equipment"].split(',')

            self.dimensions = (df["height"], df["width"], df["depth"])

        elif isinstance(df, dict):
            self.__uid = df["_Sensor__uid"]
            self.personnel = df["personnel"]
            self.equipment = df["equipment"]
            self.dimensions = df["dimensions"]
        else:
            raise Exception("Invalid Input type")

        #self.__uid = df["_Sensor__uid"]

        # assert imported data objects are of the correct type
        if not isinstance(df["name"], str):
            raise Exception("Invalid Input type")

        self.name = df["name"]
        self.method_of_sensing = df["method_of_sensing"]
        self.automation = df["automation"]
        self.sensed_variables = df["sensed_variables"]
        self.acquisition_cost = df["acquisition_cost"]
        self.replacement_cost = df["replacement_cost"]
        self.operational_cost_ph = df["operational_cost_ph"]
        self.testing_cost_ph = df["testing_cost_ph"]
        self.false_alarm_cost = df["false_alarm_cost"]
        self.detection_cost = df["detection_cost"]
        self.mttf = df["mttf"]
        self.mttr = df["mttr"]
        self.sampling_interval = df["sampling_interval"]
        self.test_duration = df["test_duration"]
        self.volume = self.dimensions[0] * self.dimensions[1] * self.dimensions[2]
        self.weight = df["weight"]
        self.error_code = df["error_code"]

        self.metric_table = self.to_metric_table()
        #self.attr_dict = dict(self.obj_2_dict())

    def to_metric_table(self):
        #tmp_uid = self.__uid
        tmp_dict = self.obj_2_dict()
        tmp_dict["height"] = self.dimensions[0]
        tmp_dict["width"] = self.dimensions[1]
        tmp_dict["depth"] = self.dimensions[2]
        tmp_dict.pop('dimensions', None)
        tmp_dict.pop('personnel', None)
        tmp_dict.pop('equipment', None)
        tmp_dict.pop('equipment', None)
        tmp_dict.pop("_Sensor__uid", None)
        df = pd.DataFrame.from_dict(tmp_dict, orient='index')
        df.rename(columns={0:self.__uid})
        return df

    def obj_2_dict(self):
        # https://stackoverflow.com/questions/61517/python-dictionary-from-an-objects-fields
        dic = dict(vars(self).copy())
        dic.pop('metric_table')
        return dic

    def obj_2_excel(self, path, sheet):
        """
        Not Needed as the database determines the object not vice versa
        :param path:
        :param sheet:
        :return:
        """
        obj_dict = self.obj_2_dict()
        #obj_df = pd.DataFrame.from_dict()


class SenseLocation:
    def __init__(self, df=None):
        self._SenseLocation__uid = None
        self.location = None
        self.sense_type = None
        self.sensor_string_list = []
        self.sensors = []

        if df is not None:
            self.dict_2_obj(df)

    def dict_2_obj(self, df):

        if isinstance(df, pd.DataFrame):
            key = list(df.keys())[0]
            self._SenseLocation__uid = key
            df = df[key]
            self.sensor_string_list = df["_Sensor__uid"].split(',')

        elif isinstance(df, dict):
            self._SenseLocation__uid = df['_SenseLocation__uid']
            self.sensor_string_list = df['_Sensor__uid']
        else:
            raise Exception("Invalid")

        self.location = df["location"]
        self.sense_type = df["sense_type"]

    def get_sensors_from_db(self, sensor_library_collection):
        for sensor_string in self.sensor_string_list:
            res = sensor_library_collection.find_one({"_Sensor__uid": sensor_string})
            self.sensors.append(Sensor(res))
            print("pause")

    def obj_2_dict(self):
        # https://stackoverflow.com/questions/61517/python-dictionary-from-an-objects-fields
        return {
            "_SenseLocation__uid": self._SenseLocation__uid,
            "location": self.location,
            "sense_type": self.sense_type,
            "_Sensor__uid": self.sensor_string_list,
        }

class FailureMode:
    def __init__(self, df=None):
        self._FailureMode__uid = None
        self.diagnostic_group = None
        self.component = None
        self.flow_property = None
        self.failure = None
        self.criticality = None
        self.cost = None
        self.must_cover = None
        self.must_detect = None
        #self.attr_dict = None
        #self.num_metrics = 6

        if df is not None:
            self.dict_2_obj(df)

    def dict_2_obj(self, df):
        if isinstance(df, pd.DataFrame):
            key = list(df.keys())[0]
            self._FailureMode__uid = key
            df = df[key]

        elif isinstance(df, dict):
            self._FailureMode__uid = df["_FailureMode__uid"]

        self.diagnostic_group = df["diagnostic_group"]
        self.component = df["component"]
        self.flow_property = df["flow_property"]
        self.criticality = df["criticality"]
        self.component = df["component"]
        self.cost = df["cost"]

        assert not (df["must_cover"] == df["must_detect"] == 1)
        self.must_cover = df["must_cover"]
        self.must_detect = df["must_detect"]

        #self.attr_dict = dict(self.obj_2_dict())

    def obj_2_dict(self):
        # https://stackoverflow.com/questions/61517/python-dictionary-from-an-objects-fields
        return dict(vars(self))


class PropagationTable:
    """
    Propagation table refers to the concrete propagation table object.
    Calculations on the self.pt are used for tracability and exportability to excel
    Calculations exist with the numpy calcs are used for speed in future genetic algorithm
    """
    def __init__(self, name, pt):
        self.name = name
        self.pt = pt
        self.fm_table = None
        self.length = self.pt.shape[0]
        self.labels = FM_LABELS
        try:
            self.numpy_pt = np.array(self.pt.drop(self.labels, axis=1))
        except:
            raise Exception('Error')

        self.s_loc_list = list(self.pt.columns)
        for el in self.labels: self.s_loc_list.remove(el)
        #failure_modes_list = get_failure_mode_list(self.pt)

        self.s_loc = {}
        for s_loc in self.s_loc_list:
            self.s_loc[s_loc] = SenseLocation()

        self.sensors = {}
        for s_loc in self.s_loc_list:
            self.sensors[s_loc] = Sensor()

        failure_modes_list = self.pt['_FailureMode__uid'] #get_failure_mode_list(self.pt)
        self.failure_modes = {}
        for fm in failure_modes_list:
            self.failure_modes[fm] = FailureMode()

        #self.amgiguety_groups = None
        self.sensor_candidates = []
        self.sensor_indexes = []

        # Metrics
        self.coverage = None
        self.detection = None
        self.undetected_criticality = None
        self.uncovered_criticality = None

        self.pt['group'] = ""

    def filter_np_pt(self, rows, cols):
        return self.numpy_pt[np.ix_(rows, cols)]

    def add_sensor_candidates(self, sensor_coll):
        cursor = sensor_coll.find({})
        for doc in cursor:
            self.sensor_candidates[doc["_Sensor__uid"]] = Sensor(doc)

    def add_sensors(self, sensor_coll, df):
        #assign sensor to each sensed_location
        for col in df:
            uid = df[col]["_Sensor__uid"]
            doc = sensor_coll.find_one({"_Sensor__uid":uid})
            self.sensors[col] = Sensor(doc)

    def add_sense_locations(self, coll):
        cols = list(self.pt.columns)
        for s_loc_uid in self.s_loc_list:
            doc = coll.find_one({"_SenseLocation__uid":s_loc_uid})
            self.s_loc[s_loc_uid] = SenseLocation(doc)

    def add_failure_modes(self, coll):
        for fm_uid in self.pt["_FailureMode__uid"]:
            doc = coll.find_one({"_FailureMode__uid":  fm_uid})
            self.failure_modes[fm_uid] = FailureMode(doc)

        self.build_fm_table()

    def calc_metrics(self):
        pass

    def calc_sum_ave(self, attribute):
        cost = 0
        for key in self.sensors:
            cost += self.sensors[key[attribute]]

        return cost, cost/len(self.sensors)

    def calc_op_cost_ph(self):
        cost = 0
        for key in self.sensors:
            cost += self.sensors[key["acquisition_cost"]]

        return cost, cost/len(self.sensors)

    def build_fm_table(self):
        i = 0
        fm_table_dict = {}
        columns = []
        keys = []
        for fm_uid in self.failure_modes:
            tmp_dic = self.failure_modes[fm_uid].obj_2_dict()

            if i == 0:
                keys = list(tmp_dic.keys())
                keys.remove('_FailureMode__uid')

            row_values = list(tmp_dic.values())
            columns.append(row_values[0])
            fm_table_dict[row_values[0]] = row_values[1:]

            i += 1

        self.fm_table = pd.DataFrame(fm_table_dict, columns=columns, index=keys)

    def calc_stats(self):
        self.calc_detection()
        self.calc_coverage()
        #Assert that the tracable calcs match the quick PythonCoverageRunConfigurationExtension
        if self.coverage != calc_coverage_np(self.numpy_pt)*100:
            raise Exception(f"{self.coverage}!={calc_coverage_np(self.numpy_pt)*100}")
        if self.detection != calc_detection_np(self.numpy_pt)*100:
            raise Exception(f"{self.detection}!={calc_detection_np(self.numpy_pt)*100}")

    def save_pt(self, writer, unique=True, s_loc=True, group=True):
        # Create pt
        cols=[]

        if unique:
            cols += self.labels
        if s_loc:
            cols += self.s_loc_list
        if group:
            cols += ['group']

        pt = self.pt[cols]
        pt.to_excel(writer, self.name + "_pt")

    def filter_propagation_table(self, name, rows, columns):
        """
        Filter the propagation table object by failure modes and columns (Or just update the baseline PT .xlsx)
        :param name:
        :param rows:
        :param columns:
        :return:
        """

    def calc_detection(self):
        """
        A failure mode is detectable if any of the sensors in its syndrome are not zero
        :return:
        """
        count = 0
        un_d_crit = 0
        for index, row in self.pt.iterrows():
            # If the sum absolute value of the propagation table row is 0, it is not detectable
            if self.pt.loc[index, self.s_loc_list].abs().sum() == 0:
                count += 1
                if self.failure_modes[self.pt.loc[index, "_FailureMode__uid"]].criticality is not None:
                    un_d_crit += self.failure_modes[self.pt.loc[index, "_FailureMode__uid"]].criticality
                    print(f"{self.pt.loc[index, '_FailureMode__uid']} criticality == None")

        self.undetected_criticality = un_d_crit
        self.detection = (1 - count / self.length) * 100

    def calc_coverage(self):
        """
        A failure mode is covered if it has a unique syndrome in the propagation table
        :return:
        """
        print(self.name)
        #pt = self.pt
        unique = {}
        un_cov = 0
        for index, row in self.pt.iterrows():
            key = str(np.array(row[self.s_loc_list]))
            if unique.get(key) is None:
                unique[key]= [index]
            else:
                unique[key].append(index)
        i = 0
        count = 0
        print(f"    {unique}")
        for key in unique:
            if len(unique[key]) == 1:
                count += 1
                for index in unique[key]:
                    print(f"        update {i}")
                    self.pt.loc[index, 'group'] = "F " + str(i)

            else:
                for index in unique[key]:
                    print(f"        update {i}")
                    self.pt.loc[index, 'group'] = "G " + str(i)
                    un_cov += self.failure_modes[self.pt.loc[index, "_FailureMode__uid"]].criticality
            i += 1

        self.uncovered_criticality = un_cov
        self.coverage = (count/self.length)*100

    def save_stats(self, writer):
        # Create Stats df:
        dic = {}
        dic["Coverage"] = self.coverage
        dic["Detection"] = self.detection
        dic["Undetected Criticality"] = self.undetected_criticality
        dic["Uncovered Criticality"] = self.uncovered_criticality
        dic["Num_sense_locations"] = len(self.s_loc_list)
        df = pd.DataFrame(dic, index=[0])
        df.to_excel(writer, self.name + "_stats")







