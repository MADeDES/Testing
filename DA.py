import pandas as pd
import numpy as np
import pymongo
import openpyxl

#Constants for proptables
from openpyxl.utils import get_column_letter

FLOW = "Flow Property"
SYMP = "Symptom"
BOTH = "Both"
COMPONENT = "Component"
FAILURE = "Failure"


def get_failure_mode_list(pt):
    """
    Get list of unique failure modes
    :return:
    """
    return pt[COMPONENT] + " " + pt[FLOW] + " " + pt[FAILURE]

def search_value_in_col_idx(ws, search_string, col_idx=1):
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx, None


def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row

def try_create_ws(filepath, sheet):
    wb = openpyxl.load_workbook(filepath, read_only=False)
    try:
        ws = wb[sheet]
    except KeyError:
        wb.create_sheet(sheet)
        ws = wb[sheet]
        # raise Exception(f"Sheet {sheet} does not exist")
    return wb, ws

def cursor_len(cursor):
    """
    Return cursor length
    :param cursor:
    :return:
    """
    i = 0
    for doc in cursor:
        i += 1
    return i

def calc_coverage_np(pt):
    (unique, counts) = np.unique(pt, axis=0, return_counts=True)
    return np.count_nonzero(counts == 1)/pt.shape[0]

def calc_detection_np(pt):
    count = 0
    for row in pt:
        if np.abs(row).sum() == 0:
            count += 1

    return 1-count/pt.shape[0]

def obj_2_collection(obj, coll):
    """
    Send object to databse in dict- > JSON -> BSON
    :param collection:
    :return:
    """
    dict_for_db = obj.obj_2_dict()

    coll.delete_one({"_Sensor__uid": dict_for_db["_Sensor__uid"]})
    #if cursor_len(cursor) == 0:
        # listing["available"] = True # Is this listing still available
    coll.insert_one(dict_for_db)

class SensorFactory:
    """
    Read in excel document, create python objects and send it to a JSON file for MongoDB, then delete the python objects
    """
    def __init__(self, path, sheetname, collection):
        df = pd.read_excel(path, sheetname, index_col=0)
        # assert that all columns of the sensor library are unique (unique uids)
        assert 1==1
        for column in df:
            sensor = Sensor(df[column].to_frame())
            obj_2_collection(sensor, collection)

class Sensor:
    """
    All attributes are private to ensure all editing occurs in the excel database
    """
    def __init__(self, df=None):
        self.__uid = None  #
        self.name = None
        self.method_of_sensing = "On-line"
        self.automation = "Automatic"
        self.sensed_variables = None
        self.acquisition_cost = 0
        self.replacement_cost = 0
        self.operational_cost_ph = 0
        self.testing_cost_ph = 0
        self.false_alarm_cost = 0
        self.detection_cost = 0
        self.mttf = None
        self.mttr = None
        self.sampling_rate = None
        self.test_duration = 0
        self.personnel = []
        self.equipment = []
        self.dimensions = (0, 0, 0)
        self.volume = None
        self.weight = 0
        self.error_code = ""

        if df is not None:
            self.dict_2_obj(df)

    def dict_2_obj(self, df):

        """
        Create object from mongodb database (dict) or from excel
        hardcode the creation of the item from the schema
        :param df:
        :return:
        """
        if isinstance(df, pd.DataFrame):
            key = list(df.keys())[0]
            self.__uid = key
            df = df[key]

            if not pd.isnull(df["personnel"]):
                self.equipment = df["equipment"].split(',')
            if not pd.isnull(df["equipment"]):
                self.equipment = df["equipment"].split(',')

            self.dimensions = (df["length"], df["width"], df["depth"])

        elif isinstance(df, dict):
            self.uid = df["_Sensor__uid"]
            self.equipment = df["equipment"]
            self.equipment = df["equipment"]
            self.dimensions = df["dimensions"]
        else:
            raise Exception("Invalid Input type")

        #self.__uid = df["_Sensor__uid"]

        # assert imported data objects are of the correct type
        if not isinstance(df["name"], str):
            raise Exception("Invalid Input type")

        self.name = df["name"]
        self.method_of_sensing = df["method_of_sensing"]
        self.automation = df["automation"]
        self.sensed_variables = df["sensed_variables"]
        self.acquisition_cost = df["acquisition_cost"]
        self.replacement_cost = df["replacement_cost"]
        self.operational_cost_ph = df["operational_cost_ph"]
        self.testing_cost_ph = df["testing_cost_ph"]
        self.false_alarm_cost = df["false_alarm_cost"]
        self.detection_cost = df["detection_cost"]
        self.mttf = df["mttf"]
        self.mttr = df["mttr"]
        self.sampling_rate = df["sampling_rate"]
        self.test_duration = df["test_duration"]
        self.volume = self.dimensions[0] * self.dimensions[1] * self.dimensions[2]
        self.weight = df["weight"]
        self.error_code = df["error_code"]

    def obj_2_dict(self):
        # https://stackoverflow.com/questions/61517/python-dictionary-from-an-objects-fields
        return vars(self)

    def obj_2_excel(self, path, sheet):
        """
        Not Needed as the database determines the object not vice versa
        :param path:
        :param sheet:
        :return:
        """
        obj_dict = self.obj_2_dict()
        #obj_df = pd.DataFrame.from_dict()

class PropagationTable:
    """
    Propagation table refers to the concrete propagation table object.
    Calculations on the self.pt are used for tracability and exportability to excel
    Calculations exist with the numpy calcs are used for speed in future genetic algorithm
    """
    def __init__(self, name, pt, parent=None, s_loc_type=None, sensors=None):
        if parent is not None:
            parent.propagation_tables[name] = self
        self.name = name
        self.pt = pt
        self.s_locs = list(self.pt.columns)
        self.labels = ['Diagnostic Group', 'Component', 'Flow Property', 'Failure']
        for el in self.labels: self.s_locs.remove(el)
        self.numpy_pt = np.array(self.pt.drop(self.labels, axis=1))

        self.s_loc_type = s_loc_type
        self.length = self.pt.shape[0]
        self.coverage = None
        self.detection = None
        self.sensors = {}
        for s_loc in self.s_locs:
            self.sensors[s_loc] = Sensor()
        self.failure_modes = get_failure_mode_list(self.pt)
        #self.amgiguety_groups = None

        self.pt['group'] = ""

    def filter_np_pt(self, rows, cols):
        return self.numpy_pt[np.ix_(rows, cols)]

    def add_sensors(self, sensor_coll, df):
        for index in df.index:
            uid = df.loc[index,"Sensor"]
            doc = sensor_coll.find_one({"_Sensor__uid":uid})
            self.sensors[df.loc[index,"Sense Location"]] = Sensor(doc)

    def calc_metrics(self):
        pass

    def calc_sum_ave(self, attribute):
        cost = 0
        for key in self.sensors:
            cost += self.sensors[key[attribute]]

        return cost, cost/len(self.sensors)

    def calc_op_cost_ph(self):
        cost = 0
        for key in self.sensors:
            cost += self.sensors[key["acquisition_cost"]]

        return cost, cost/len(self.sensors)


    def calc_stats(self):
        self.calc_detection()
        self.calc_coverage()
        #Assert that the tracable calcs match the quick PythonCoverageRunConfigurationExtension
        if self.coverage != calc_coverage_np(self.numpy_pt)*100:
            raise Exception(f"{self.coverage}!={calc_coverage_np(self.numpy_pt)*100}")
        if self.detection != calc_detection_np(self.numpy_pt)*100:
            raise Exception(f"{self.detection}!={calc_detection_np(self.numpy_pt)*100}")

    def filter_propagation_table(self, name, rows, columns):
        """
        Filter the propagation table object by failure modes and columns (Or just update the baseline PT .xlsx)
        :param name:
        :param rows:
        :param columns:
        :return:
        """


    def calc_detection(self):
        """
        A failure mode is detectable if any of the sensors in its syndrome are not zero
        :param sensors:
        :return:
        """

        count = 0
        for index, row in self.pt.iterrows():
            # If the sum absolute value of the propagation table row is 0, it is not detectable
            if self.pt.loc[index, self.s_locs].abs().sum() == 0:
                count += 1

        self.detection = (1 - count / self.length) * 100

    def calc_coverage(self):
        """
        A failure mode is covered if it has a unique syndrome in the propagation table
        :param sensors:
        :return:
        """
        print(self.name)
        #pt = self.pt
        unique = {}
        for index, row in self.pt.iterrows():
            key = str(np.array(row[self.s_locs]))
            if unique.get(key) is None:
                unique[key]= [index]
            else:
                unique[key].append(index)
        i = 0
        count = 0
        print(f"    {unique}")
        for key in unique:
            if len(unique[key]) == 1:
                count += 1
                for index in unique[key]:
                    print(f"        update {i}")
                    self.pt.loc[index, 'group'] = "F " + str(i)

            else:
                for index in unique[key]:
                    print(f"        update {i}")
                    self.pt.loc[index, 'group'] = "G " + str(i)
            i += 1

        self.coverage = (count/self.length)*100


    def save_fm(self, filepath, sheet):
        wb, ws = try_create_ws(filepath, sheet)

        ws["A" + str(1)] = "Failure Mode"
        i=2
        for header in self.labels:
            letter = openpyxl.utils.get_column_letter(i)
            ws[letter + str(1)] = header
            i+=1

        letter = openpyxl.utils.get_column_letter(i)
        ws[letter + str(1)] = "Mask"
        letter = openpyxl.utils.get_column_letter(i + 1)
        ws[letter + str(1)] = "Must Cover / Detect"  # "","Cover","Detect"
        letter = openpyxl.utils.get_column_letter(i + 2)
        ws[letter + str(1)] = "Criticality"  # "","Cover","Detect"

        tmp_dict = {}
        #pop sensor to ensure deltion is working (commented out for normal operation)
        #tmp = self.sensors.pop()
        #print(f"del {tmp}")

        # append new sensors to the list
        for i, fm in enumerate(self.failure_modes):
            tmp_dict[fm] = ""
            col_idx, row = search_value_in_col_idx(ws, fm, col_idx=0)
            if row == None:
                ws.append([fm] + list(self.pt.iloc[i,0:4].values)) # + ["", f"=IF(OR(F{i}='', F{i}='Cover', F{i}='Detect'), TRUE, FALSE)"]

        # Delete rows that no longer are in the sensor set
        if len(self.failure_modes)<ws.max_row-1: #one row is a header
            idx = 2
            for row in ws.iter_rows(min_row=idx):
                if tmp_dict.get(row[0].value) is None:
                    print(f"deleted {row[0].value} row {idx}")
                    ws.delete_rows(idx=idx)

                idx+=1

        wb.save(filepath)

    def save_sense(self, filepath, sheet):
        wb, ws = try_create_ws(filepath, sheet)

        ws["A" + str(1)] = "Sense Location"
        ws["B" + str(1)] = "Sense Type"
        ws["C" + str(1)] = "Filter"
        ws["D" + str(1)] = "Sensor"

        tmp_dict = {}
        #pop sensor to ensure deltion is working (commented out for normal operation)
        #tmp = self.sensors.pop()
        #print(f"del {tmp}")

        # append new sensors to the list
        for i, s_loc_type in enumerate(self.s_locs):
            tmp_dict[s_loc_type]=""
            col_idx, row = search_value_in_col_idx(ws, s_loc_type, col_idx=0)
            if row == None:
                ws.append([self.s_locs[i], self.s_loc_type[i]])

        # Delete rows that no longer are in the sensor set
        if len(self.s_locs)<ws.max_row-1: #one row is a header
            idx = 2
            for row in ws.iter_rows(min_row=idx):
                if tmp_dict.get(row[0].value) is None:
                    print(f"deleted {row[0].value} row {idx}")
                    ws.delete_rows(idx=idx)

                idx+=1

        wb.save(filepath)

    def save_pt(self, writer, unique=True, s_loc=True, group=True):
        # Create pt
        cols=[]

        if unique:
            cols += self.labels
        if s_loc:
            cols += self.s_locs
        if group:
            cols += ['group']

        pt = self.pt[cols]
        pt.to_excel(writer, self.name + "_pt")

    def save_stats(self, writer):

        # Create Stats df:
        dic = {}
        dic["Coverage"] = self.coverage
        dic["Detection"] = self.detection
        dic["Num_sense_locations"] = len(self.s_locs)
        df = pd.DataFrame(dic, index=[0])
        df.to_excel(writer, self.name + "_stats")

class Diag_Analysis:
    def __init__(self, filename, name):
        self.filename = filename
        self.filename_xlsx = filename + ".xlsx"
        self.name = name
        self.labels = ['Diagnostic Group', 'Component', 'Flow Property', 'Failure']

        # get PT and Merge
        pt_flow = pd.read_excel(self.filename_xlsx, sheet_name=FLOW).replace(np.nan, 0)
        pt_symp = pd.read_excel(self.filename_xlsx, sheet_name=SYMP).replace(np.nan, 0)
        assert pt_flow['Diagnostic Group'].equals(pt_symp['Diagnostic Group'])
        assert pt_flow['Component'].equals(pt_symp['Component'])
        assert pt_flow['Flow Property'].equals(pt_symp['Flow Property'])
        assert pt_flow['Failure'].equals(pt_symp['Failure'])

        # Get sensor locations and merge
        symp_cols = list(pt_symp.columns)
        flow_cols = list(pt_flow.columns)
        for el in self.labels: symp_cols.remove(el)
        for el in self.labels: flow_cols.remove(el)
        both_cols = flow_cols + symp_cols

        pt_both = pd.concat([pt_flow, pt_symp[symp_cols]], axis=1)
        # pt_both.reset_index(inplace=True, drop=True)
        assert pt_both['Diagnostic Group'].equals(pt_symp['Diagnostic Group'])
        assert pt_both['Component'].equals(pt_symp['Component'])
        assert pt_both['Flow Property'].equals(pt_symp['Flow Property'])
        assert pt_both['Failure'].equals(pt_symp['Failure'])

        # Create PT objects
        symp_s_loc_types = [SYMP]*len(symp_cols)
        flow_s_loc_types = [FLOW]*len(flow_cols)
        both_s_loc_types = flow_s_loc_types + symp_s_loc_types
        self.propagation_tables = {}
        self.symptoms = PropagationTable(SYMP, pt_symp, parent=self, s_loc_type=symp_s_loc_types)
        self.flow = PropagationTable(FLOW, pt_flow, parent=self,  s_loc_type=flow_s_loc_types)
        self.both = PropagationTable(BOTH, pt_both, parent=self,  s_loc_type=both_s_loc_types)

        # self.symptoms.do_calcs()
        # self.flow.do_calcs()
        # self.both.do_calcs()
        print("here")

    def add_pt(self, name):
        PropagationTable(name, self.both.pt.copy(), parent=self)

    def calc_stats(self):
        for pt in self.propagation_tables:
            self.propagation_tables[pt].calc_stats()

    def save_stats(self):
        with pd.ExcelWriter(self.filename+"_out"+".xlsx") as writer:
            for pt in self.propagation_tables:
                self.propagation_tables[pt].save_pt(writer)
                self.propagation_tables[pt].save_stats(writer)

    def include(self):
        pass



# "C:/Users/61435/OneDrive - PHM Technology/PHM Tech/Research/Testability/PHM_Module"
pt_path = "C:/Users/61435/OneDrive - PHM Technology/PHM Tech/Research/Testability/PHM_Module/PT/"
pt_filename = "ProptableMini"
sensor_path = "C:/Users/61435/OneDrive - PHM Technology/PHM Tech/Research/Testability/PHM_Module/"
sensor_filename = "Excel_Gui.xlsx"

# Create open_pyxl doc

# Pymongo SetUp Database
client = pymongo.MongoClient(host="localhost", port=27017)
MADe_db = client.MADe
sensor_library_collection = MADe_db.Sensor_Library

SensorFactory(sensor_path+sensor_filename, "sensor_library", sensor_library_collection)
base_da = Diag_Analysis(pt_path + pt_filename, "Base")
# S1 = ['C1 (Pneumatic - Mass flow rate)', 'C2 (Pneumatic - Mass flow rate)', 'Combustion System (Continuous - Data)', 'Engine (Discrete - Data)', 'Engine (Discrete - Data).1', 'Other G1']
# A3_S1_tD = ['C1 (Pneumatic - Mass flow rate)','C1 (Pneumatic - Mass flow rate).1','C2 (Pneumatic - Mass flow rate)','CV1 (Liquid - Flow rate)','CV1 (Liquid - Contamination)','Combustion System (Mechanical - rotational - Angular velocity)','Combustion System (Pneumatic - Mass flow rate)','Combustion System (Mechanical - rotational - Angular velocity)','O Pump (Gas - Static pressure)','Engine (Discrete - Data)','Engine (Discrete - Data).1','Engine (Electrical - Voltage)','CV2 (Gas - Mass flow rate)', 'Exhaust Nozzle (Continuous - Data)', 'Exhaust Nozzle (Pneumatic - Mass flow rate)', 'Vibration G1',  'Change in Behaviour']
# A3_S1_tD.sort()
# A4_S3 = ['Exhaust Nozzle (Pneumatic - Mass flow rate)', 'Vibration G1', 'C1 (Pneumatic - Mass flow rate).1']
# A4_S6 = ['O Pump (Gas - Static pressure)', 'Vibration G1', 'C1 (Pneumatic - Mass flow rate).1']

# base_da.add_pt('S1', S1)
# base_da.add_pt('A3 - S1 - test Diagnostic', A3_S1_tD)
# base_da.add_pt('A4 - S3 - John Yoo', A4_S3)
# base_da.add_pt('A4 - S6 - John Yoo', A4_S6)
# base_da.add_pt('S1', S1)

base_da.calc_stats()
base_da.save_stats()

base_da.both.save_sense(sensor_path+sensor_filename, "sense_locations")
base_da.both.save_fm(sensor_path+sensor_filename, "failure_modes")
# with pd.ExcelWriter(sensor_path+sensor_filename) as writer:
#     base_da.both.save_pt(writer)



# Calc metrics from sensor set
df = pd.read_excel(sensor_path+sensor_filename, "sense_locations")
sensed_locs = df["Sense Location"][df["Sensor"] != ""]
new_pt = PropagationTable("SensorSetMetrics", base_da.both.pt[base_da.both.labels+list(sensed_locs)])
new_pt.calc_stats()
new_pt.add_sensors(sensor_library_collection, df)
# Read in excel file with failure modes and sensors

# Ca

